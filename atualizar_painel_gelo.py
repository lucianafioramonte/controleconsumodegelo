# -*- coding: utf-8 -*-
"""
Atualiza o painel Torre de Controle - Consumo de Gelo.

Le a planilha RESUMO USINA CONCRETO.xlsx (aba RESUMO), recalcula os
indicadores de consumo de gelo e gera o arquivo Torre_Controle_Gelo.html
atualizado.

Uso:
    python atualizar_painel_gelo.py

Ajuste os caminhos abaixo (SOURCE_XLSX / OUTPUT_DIR) se as pastas mudarem.
"""

import openpyxl
import datetime
import json
import collections
import statistics
import sys
import os

# ------------------------------------------------------------------
# CAMINHOS - ajuste aqui se necessario
# ------------------------------------------------------------------
SOURCE_XLSX = r"C:\Users\lucianaboas\OneDrive - Construtora Tripoloni\Planejamento\08 - Controles\06. Controle da Usina de Concreto\2026\RESUMO USINA CONCRETO.xlsx"
OUTPUT_DIR = r"C:\Users\lucianaboas\OneDrive - Construtora Tripoloni\Documentos\Controle de Gelo"
OUTPUT_HTML = os.path.join(OUTPUT_DIR, "Torre_Controle_Gelo.html")

# o template.html deve estar na mesma pasta deste script
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_HTML = os.path.join(SCRIPT_DIR, "template.html")

SHEET_NAME = "RESUMO"
FIRST_DATA_ROW = 9
SACO_KG = 20  # peso de cada saco de gelo


# ------------------------------------------------------------------
# CLASSIFICACAO
# ------------------------------------------------------------------
def classify_elemento(destino):
    if not destino:
        return "Não informado"
    d = str(destino).upper()
    rules = [
        ("Aduela", ["ADUELA"]),
        ("Viaduto", ["VIADUTO", "VIADULTO"]),
        ("Travessa", ["TRAVESSA", "TRAVASSA"]),
        ("Transversina", ["TRANSVERSINA"]),
        ("Viga", ["VIGA"]),
        ("Laje", ["LAJE", "LAGE"]),
        ("Estaca", ["ESTACA", "ESTAC"]),
        ("Pilar", ["PILAR"]),
        ("Galeria", ["GALERIA", "GALARIA"]),
        ("Dissipador", ["DISSIPADOR"]),
        ("Ala/Cortina", ["ALA", "CORTINA"]),
        ("Guarda-rodas", ["GUARDA"]),
        ("Canaleta", ["CANALETA", "CANELETA"]),
        ("Caixa/Drenagem", ["CAIXA", "DRENAGEM", "BUEIRO"]),
        ("Berço", ["BERÇO", "BERCO"]),
        ("Bloco de Fundação", ["BLOCO"]),
        ("Magro", ["MAGRO"]),
        ("Calçada/Meio-fio", ["CALÇADA", "CALCADA", "MEIO FIO", "MEIO-FIO"]),
        ("Escada", ["ESCADA"]),
        ("Tubulação", ["TUBULA"]),
    ]
    for name, kws in rules:
        for kw in kws:
            if kw in d:
                return name
    return "Outros"


def classify_frente(obs, destino):
    if obs:
        o = str(obs).upper()
        if "TELES PIRES" in o or "PONTE" in o or "TRELI" in o:
            return "Ponte Teles Pires"
        if "PRIMAVERA" in o:
            return "Diamante Primavera"
        if "SORRISO" in o:
            return "Diamante Sorriso"
        if "GALERIA" in o:
            return "Galeria"
        if "CANTEIRO" in o:
            return "Canteiro / Pátio de Pré-moldados"
        if "LAVAGEM" in o or "DESCART" in o:
            return "Outros/Observação"
        return "Outros/Observação"
    return "Sem frente informada"


# ------------------------------------------------------------------
# EXTRACAO
# ------------------------------------------------------------------
def extract_rows(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[SHEET_NAME]
    rows = []
    for i, row in enumerate(
        ws.iter_rows(min_row=FIRST_DATA_ROW, max_row=ws.max_row, values_only=True),
        start=FIRST_DATA_ROW,
    ):
        (date, tick, codigo, hent, hsai, mpa, traco, destino, obs,
         km, vol, tusina, tcampo, gelo_sacos, dmt, volacum) = row
        if date is None or not isinstance(date, datetime.datetime):
            continue
        rows.append({
            "row": i,
            "date": date.strftime("%Y-%m-%d"),
            "codigo": codigo,
            "destino": destino,
            "obs": obs,
            "volume": vol if isinstance(vol, (int, float)) else None,
            "temp_usina": tusina if isinstance(tusina, (int, float)) else None,
            "temp_campo": tcampo if isinstance(tcampo, (int, float)) else None,
            "gelo_sacos": gelo_sacos if isinstance(gelo_sacos, (int, float)) else None,
            "elemento": classify_elemento(destino),
            "frente": classify_frente(obs, destino),
        })
    return rows


# ------------------------------------------------------------------
# AGREGACAO
# ------------------------------------------------------------------
def build_aggregates(rows):
    by_date_temp = collections.defaultdict(list)
    for r in rows:
        if r["temp_usina"] is not None:
            by_date_temp[r["date"]].append(r["temp_usina"])
    daily_temp = []
    for d in sorted(by_date_temp.keys()):
        vals = by_date_temp[d]
        daily_temp.append({
            "date": d, "avg": round(statistics.mean(vals), 1),
            "min": min(vals), "max": max(vals), "n": len(vals),
        })

    ice = [r for r in rows if r["gelo_sacos"] is not None]
    for r in ice:
        r["gelo_kg"] = r["gelo_sacos"] * SACO_KG
        r["kg_m3"] = round(r["gelo_kg"] / r["volume"], 1) if r["volume"] else None
        r["delta_temp"] = (
            round(r["temp_usina"] - r["temp_campo"], 1)
            if (r["temp_usina"] is not None and r["temp_campo"] is not None)
            else None
        )

    by_date_gelo = collections.defaultdict(lambda: {"sacos": 0, "kg": 0, "volume": 0.0, "n": 0})
    for r in ice:
        v = by_date_gelo[r["date"]]
        v["sacos"] += r["gelo_sacos"]
        v["kg"] += r["gelo_kg"]
        v["n"] += 1
        if r["volume"]:
            v["volume"] += r["volume"]
    daily_gelo = []
    for d in sorted(by_date_gelo.keys()):
        v = by_date_gelo[d]
        kg_m3 = round(v["kg"] / v["volume"], 1) if v["volume"] else None
        daily_gelo.append({
            "date": d, "sacos": v["sacos"], "kg": v["kg"],
            "volume": round(v["volume"], 1), "kg_m3": kg_m3, "n": v["n"],
        })

    by_elem = collections.defaultdict(lambda: {"sacos": 0, "kg": 0, "volume": 0.0, "n": 0, "n_vol": 0})
    for r in ice:
        v = by_elem[r["elemento"]]
        v["sacos"] += r["gelo_sacos"]
        v["kg"] += r["gelo_kg"]
        v["n"] += 1
        if r["volume"]:
            v["volume"] += r["volume"]
            v["n_vol"] += 1
    elem_list = []
    for e, v in sorted(by_elem.items(), key=lambda x: -x[1]["kg"]):
        kg_m3_medio = round(v["kg"] / v["volume"], 1) if v["volume"] else None
        vol_medio = round(v["volume"] / v["n_vol"], 1) if v["n_vol"] else None
        elem_list.append({
            "elemento": e, "sacos": v["sacos"], "kg": v["kg"],
            "volume": round(v["volume"], 1), "n": v["n"],
            "kg_m3_medio": kg_m3_medio, "vol_medio_viagem": vol_medio,
        })

    by_frente = collections.defaultdict(lambda: {"sacos": 0, "kg": 0, "n": 0})
    for r in ice:
        v = by_frente[r["frente"]]
        v["sacos"] += r["gelo_sacos"]
        v["kg"] += r["gelo_kg"]
        v["n"] += 1
    frente_list = [
        {"frente": f, "sacos": v["sacos"], "kg": v["kg"], "n": v["n"]}
        for f, v in sorted(by_frente.items(), key=lambda x: -x[1]["kg"])
    ]

    scatter = [
        {"temp": r["temp_usina"], "kg_m3": r["kg_m3"], "date": r["date"], "elemento": r["elemento"]}
        for r in ice if r["temp_usina"] is not None and r["kg_m3"] is not None
    ]

    perda_pairs = [
        {"date": r["date"], "elemento": r["elemento"], "temp_usina": r["temp_usina"],
         "temp_campo": r["temp_campo"], "delta": r["delta_temp"]}
        for r in rows if r["temp_usina"] is not None and r["temp_campo"] is not None
    ]

    total_sacos = sum(r["gelo_sacos"] for r in ice)
    total_kg = total_sacos * SACO_KG
    total_vol_com_gelo = sum(r["volume"] for r in ice if r["volume"])
    kg_m3_medio = round(total_kg / total_vol_com_gelo, 1) if total_vol_com_gelo else None
    temp_usina_medio_geral = round(
        statistics.mean([r["temp_usina"] for r in rows if r["temp_usina"] is not None]), 1
    ) if any(r["temp_usina"] is not None for r in rows) else None
    temp_usina_medio_periodo_gelo = round(
        statistics.mean([r["temp_usina"] for r in ice if r["temp_usina"] is not None]), 1
    ) if any(r["temp_usina"] is not None for r in ice) else None
    n_temp_campo = sum(1 for r in rows if r["temp_campo"] is not None)

    kpis = {
        "total_sacos": total_sacos,
        "total_kg": total_kg,
        "total_vol_com_gelo": round(total_vol_com_gelo, 1),
        "kg_m3_medio": kg_m3_medio,
        "temp_usina_medio_geral": temp_usina_medio_geral,
        "temp_usina_medio_periodo_gelo": temp_usina_medio_periodo_gelo,
        "n_temp_campo": n_temp_campo,
        "n_viagens_gelo": len(ice),
        "periodo_temp": [daily_temp[0]["date"], daily_temp[-1]["date"]] if daily_temp else None,
        "periodo_gelo": [daily_gelo[0]["date"], daily_gelo[-1]["date"]] if daily_gelo else None,
    }

    return {
        "daily_temp": daily_temp,
        "daily_gelo": daily_gelo,
        "by_elemento": elem_list,
        "by_frente": frente_list,
        "scatter": scatter,
        "perda_pairs": perda_pairs,
        "kpis": kpis,
        "ice_trips": ice,
    }


# ------------------------------------------------------------------
# GERACAO DO HTML
# ------------------------------------------------------------------
def build_html(aggregates):
    with open(TEMPLATE_HTML, encoding="utf-8") as f:
        template = f.read()
    data_json = json.dumps(aggregates, ensure_ascii=False)
    if "__DATA_JSON__" not in template:
        raise RuntimeError("template.html nao contem o marcador __DATA_JSON__")
    return template.replace("__DATA_JSON__", data_json)


def main():
    print(f"[{datetime.datetime.now()}] Lendo planilha: {SOURCE_XLSX}")
    if not os.path.exists(SOURCE_XLSX):
        print("ERRO: planilha nao encontrada no caminho configurado.")
        sys.exit(1)

    rows = extract_rows(SOURCE_XLSX)
    print(f"  -> {len(rows)} linhas com data lidas")

    aggregates = build_aggregates(rows)
    print(f"  -> {aggregates['kpis']['n_viagens_gelo']} viagens com registro de gelo")

    html = build_html(aggregates)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    with open(OUTPUT_HTML, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"[{datetime.datetime.now()}] Painel atualizado: {OUTPUT_HTML}")


if __name__ == "__main__":
    main()
